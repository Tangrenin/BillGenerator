import shutil
from pathlib import Path
from config.infos_script import path_to_data
from scripts.data_utils import client_extraction
from openpyxl import load_workbook

"""
The main job of this script is the generation of the data directory, and both the data sheets and the student database
 inside ready to be used
"""


def data_templates_generation(year):
    months = {1: 'Janvier',
              2: 'Fevrier',
              3: 'Mars',
              4: 'Avril',
              5: 'Mai',
              6: 'Juin',
              7: 'Juillet',
              8: 'Aout',
              9: 'Septembre',
              10: 'Octobre',
              11: 'Novembre',
              12: 'Decembre'}
    # Check if data directory exists, else create it.
    Path(path_to_data).mkdir(parents=True, exist_ok=True)
    # Check if current year folder exists in the data folder, else create it
    Path(f'{path_to_data}{year}').mkdir(parents=True, exist_ok=True)

    # Check if student database exists, if not create it
    if not Path(f"{path_to_data}Infos_élèves.ods").is_file():
        shutil.copy("templates/Infos_élèves.ods", f"{path_to_data}Infos_élèves.ods")
        # And create all missing month sheets
        for n, month in months.items():
            file_path_no_ext = f'{path_to_data}{year}/{n}-Cours{month}'
            if (not Path(f'{file_path_no_ext}.ods').is_file()) and (not Path(f'{file_path_no_ext}.xlsx').is_file()):
                shutil.copy('templates/0-CoursMois.xlsx', f'{file_path_no_ext}.xlsx')
    else:
        clients = client_extraction()
        # creation and flattening of the list of all attached students for every clients
        all_students = [student for students in clients["Elèves rattachés"] for student in students.split(' ; ')]

        # Then adding every student in the database into the data sheets
        book = load_workbook('templates/0-CoursMois.xlsx')
        sheet = book.active
        for col, student in enumerate(all_students, start=2):
            sheet.cell(row=1, column=col).value = student
        # And creating the data sheets in the relevant folders
        for n, month in months.items():
            file_path_no_ext = f'{path_to_data}{year}/{n}-Cours{month}'
            if (not Path(f'{file_path_no_ext}.ods').is_file()) and (not Path(f'{file_path_no_ext}.xlsx').is_file()):
                book.save(f'{file_path_no_ext}.xlsx')
